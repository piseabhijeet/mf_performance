#!/usr/bin/env python3
"""
mf_vs_nifty_analysis.py

- Uses https://api.mfapi.in/mf to get the full list of schemes (schemeCode & schemeName)
- Matches user-supplied (possibly incomplete) fund names to best scheme using fuzzy matching
- Fetches NAV history for matched schemes, filters last DAYS
- Fetches NIFTY (^NSEI) daily Open/Close via yfinance for same date range
- For each trading date:
    - NIFTY Start = Open, NIFTY End = Close, NIFTY % Change = (Close-Open)/Open*100
    - Fund Start = previous available NAV (previous trading day's close), Fund End = today's NAV
      (mfapi provides only end-of-day NAV; therefore "start" is previous NAV)
    - Fund % Change = (Fund End - Fund Start)/Fund Start*100 (same as pct_change)
- Writes one Excel sheet per fund (date, Fund Start, Fund End, Fund % Change, NIFTY Start, NIFTY End, NIFTY % Change)
- Summary sheet (first sheet) with overall metrics (unchanged)
- Applies simple color fills to Summary (High tolerance = green, Low = red)
"""

import requests
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime, timedelta
from difflib import SequenceMatcher, get_close_matches
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------- USER EDITABLE ----------
FUND_QUERIES = [
    "Parag Parikh ELSS Tax Saver Fund - Direct Growth",
    "Bajaj Finserv ELSS Tax Saver Fund - Direct  - Growth",
    "Mirae Asset ELSS Tax Saver - Direct plan - Growth",
    "Canara Robeco ELSS Tax Saver - Direct plan - Growth Option",
    "Bajaj Finserv Gilt Fund - Direct - Growth",
    "Zerodha Gold ETF FoF",
    "Motilal Oswal Small Cap Fund - Direct - Growth",
    "Axis Small Cap Fund - Direct plan - Growth",
    "Kotak Nifty Next 50 Index  - Direct Plan - Growth Option",
    "Aditya Birla Sun Life ELSS Tax Saver Fund - Growth - Direct plan",
    "DSP Natural Resources and New Energy Fund - Direct plan - Growth",
    "Tata Resources & Energy Fund- Direct plan-Growth",
    "Bajaj Finserv Healthcare Fund - Direct - Growth",
    "DSP ELSS Tax Saver Fund - Direct plan - Growth",
    "DSP Nifty Next 50 Index Fund - Direct plan - Growth",
    "DSP India T.I.G.E.R. - Direct plan - Growth",
    "PGIM India Flexi Cap - Direct plan - Growth Option",
    "Mirae Asset Small Cap - Direct plan - Growth",
    "ICICI Prudential Nifty Smallcap 250 Index - Direct plan - Growth",
    "DSP Business Cycle Fund - Direct - Growth",
    "SBI MNC Fund - Direct plan - Growth"
]

OUTPUT_FILE = f"mf_vs_nifty_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
DAYS = 30  # user-specified lookback in days
MF_LIST_URL = "https://api.mfapi.in/mf"
MF_DETAIL_URL = "https://api.mfapi.in/mf/{}"  # format with scheme code
NIFTY_TICKER = "^NSEI"
REQUEST_TIMEOUT = 20

# ---------- UTILITIES ----------
def sanitize_sheet_name(name: str, max_len: int = 31) -> str:
    # remove characters invalid in Excel sheet names
    name = re.sub(r"[:\\/?*\[\]]", "_", name)
    if len(name) > max_len:
        return name[:max_len]
    return name

def similar(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

# ---------- FETCH FUND LIST ----------
def fetch_mf_list():
    print("Fetching full fund list from mfapi...")
    r = requests.get(MF_LIST_URL, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return r.json()  # list of dicts with 'schemeCode' and 'schemeName'

# ---------- MATCH SCHEME ----------
def find_best_scheme(fund_list, query):
    query_l = query.lower().strip()
    # Try substring match first (fast and often exact)
    substring_matches = [f for f in fund_list if query_l in f["schemeName"].lower()]
    if len(substring_matches) == 1:
        return substring_matches[0], 1.0  # exact-ish match
    elif len(substring_matches) > 1:
        # choose the longest matching name (heuristic) or best similarity
        best = max(substring_matches, key=lambda x: similar(query, x["schemeName"]))
        return best, similar(query, best["schemeName"])

    # otherwise fallback to fuzzy best match (SequenceMatcher)
    best = None
    best_score = 0.0
    for f in fund_list:
        score = similar(query, f["schemeName"])
        if score > best_score:
            best_score = score
            best = f
    return best, best_score

# ---------- FETCH MF DETAILS ----------
def fetch_mf_details(scheme_code):
    url = MF_DETAIL_URL.format(scheme_code)
    r = requests.get(url, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return r.json()

# ---------- FETCH NIFTY (Open & Close) ----------
def fetch_nifty(start_date, end_date):
    print(f"Fetching NIFTY ({NIFTY_TICKER}) from {start_date.date()} to {end_date.date()} ...")
    df = yf.download(NIFTY_TICKER, start=start_date, end=end_date + timedelta(days=1), progress=False)
    # Flatten columns if MultiIndex
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = ['_'.join(map(str, col)).strip() for col in df.columns.values]
    # find Open and Close columns robustly
    open_col = next((c for c in df.columns if c.lower().startswith("open")), None)
    close_col = next((c for c in df.columns if "close" in c.lower()), None)
    if open_col is None or close_col is None:
        raise RuntimeError("Could not find Open/Close columns for NIFTY from yfinance output.")
    df = df[[open_col, close_col]].reset_index()
    df.columns = ["date", "Open", "Close"]
    df['date'] = pd.to_datetime(df['date']).dt.normalize()
    df['Open'] = pd.to_numeric(df['Open'], errors='coerce')
    df['Close'] = pd.to_numeric(df['Close'], errors='coerce')
    df = df.dropna(subset=['Open', 'Close']).sort_values('date').reset_index(drop=True)
    # NIFTY percent change for the day = (Close - Open) / Open * 100
    df['% Change'] = (df['Close'] - df['Open']) / df['Open'] * 100
    return df

# ---------- PROCESS EACH FUND & SUMMARY METRICS ----------
def analyze_funds(fund_queries):
    mf_list = fetch_mf_list()
    start_date = (datetime.today() - timedelta(days=DAYS)).replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    nifty_df = fetch_nifty(start_date, end_date)

    writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')
    summary_rows = []

    for q in fund_queries:
        print(f"\n=== Processing query: '{q}' ===")
        best, score = find_best_scheme(mf_list, q)
        if not best:
            print(f"[!] No match found for: {q}")
            continue
        scheme_code = best['schemeCode']
        scheme_name = best['schemeName']
        print(f"[i] Best match: '{scheme_name}' (schemeCode={scheme_code}), score={score:.3f}")

        try:
            details = fetch_mf_details(scheme_code)
            meta = details.get('meta', {})
            data = details.get('data', [])
            if not data:
                print(f"[!] No NAV data for scheme {scheme_name} ({scheme_code})")
                continue
            df = pd.DataFrame(data)
            # parse and clean
            df['date'] = pd.to_datetime(df['date'], format="%d-%m-%Y", errors='coerce')
            df['NAV'] = pd.to_numeric(df['nav'], errors='coerce')
            df = df[['date', 'NAV']].dropna().sort_values('date').reset_index(drop=True)
            # filter last DAYS days
            df = df[df['date'] >= start_date].copy()
            if df.empty:
                print(f"[!] No NAV points in last {DAYS} days for {scheme_name}")
                continue
            df['date'] = df['date'].dt.normalize()

            # Fund start is previous available NAV (mfapi gives only EOD NAV), fund end is current NAV
            df['Fund Start'] = df['NAV'].shift(1)
            df['Fund End'] = df['NAV']
            # daily percent change for fund (based on previous NAV)
            df['Fund % Change'] = (df['Fund End'] - df['Fund Start']) / df['Fund Start'] * 100

            # Merge with NIFTY on trading days (inner join)
            # nifty_df has columns: date, Open, Close, % Change (Open→Close pct)
            merged = pd.merge(df[['date', 'Fund Start', 'Fund End', 'Fund % Change']], 
                              nifty_df[['date', 'Open', 'Close', '% Change']],
                              on='date', how='inner')

            # Rename NIFTY columns for clarity
            merged.rename(columns={'Open': 'NIFTY Start', 'Close': 'NIFTY End', '% Change': 'NIFTY % Change'}, inplace=True)

            # drop rows where either change is NaN (e.g., first fund row has no Fund Start)
            merged = merged.dropna(subset=['Fund % Change', 'NIFTY % Change'])
            if merged.empty:
                print(f"[!] After aligning with NIFTY, no overlapping trading days (with valid changes) for {scheme_name}")
                continue

            # compute metrics (preserve original logic but operate on new column names)
            corr = merged['Fund % Change'].corr(merged['NIFTY % Change'])
            signs_equal = np.sign(merged['Fund % Change']) == np.sign(merged['NIFTY % Change'])
            with_pct = (signs_equal.sum() / len(merged)) * 100

            up_mask = merged['NIFTY % Change'] > 0
            down_mask = merged['NIFTY % Change'] < 0
            avg_nifty_up = merged.loc[up_mask, 'NIFTY % Change'].mean() if up_mask.any() else np.nan
            avg_fund_up = merged.loc[up_mask, 'Fund % Change'].mean() if up_mask.any() else np.nan
            avg_nifty_down = merged.loc[down_mask, 'NIFTY % Change'].mean() if down_mask.any() else np.nan
            avg_fund_down = merged.loc[down_mask, 'Fund % Change'].mean() if down_mask.any() else np.nan

            up_capture = (avg_fund_up / avg_nifty_up * 100) if (pd.notna(avg_fund_up) and pd.notna(avg_nifty_up) and avg_nifty_up != 0) else np.nan
            down_capture = (abs(avg_fund_down) / abs(avg_nifty_down) * 100) if (pd.notna(avg_fund_down) and pd.notna(avg_nifty_down) and avg_nifty_down != 0) else np.nan

            if pd.notna(corr):
                if corr >= 0.6:
                    behavior = "With Market"
                elif corr <= -0.2:
                    behavior = "Against Market"
                else:
                    behavior = "Low/Neutral Corr"
            else:
                behavior = "Insufficient Data"

            if pd.notna(up_capture) and pd.notna(down_capture):
                if (up_capture > 100) and (down_capture < 100):
                    tolerance = "High"
                elif (up_capture < 90) and (down_capture > 120):
                    tolerance = "Low"
                else:
                    tolerance = "Medium"
            else:
                tolerance = "Unknown"

            # store summary row (fields kept as original)
            summary_rows.append({
                "Query": q,
                "Matched Scheme": scheme_name,
                "Scheme Code": scheme_code,
                "Fund House": meta.get('fund_house', ''),
                "Data Points": len(merged),
                "Correlation": round(corr, 3) if pd.notna(corr) else None,
                "With Market %": round(with_pct, 1),
                "Avg Fund Return (%)": round(merged['Fund % Change'].mean(), 4),
                "Avg Nifty Return (%)": round(merged['NIFTY % Change'].mean(), 4),
                "Up Capture (%)": round(up_capture, 1) if pd.notna(up_capture) else None,
                "Down Capture (%)": round(down_capture, 1) if pd.notna(down_capture) else None,
                "Behavior": behavior,
                "Market Tolerance": tolerance
            })

            # Save merged data to a sheet named with scheme code and short scheme name
            sheet_name = sanitize_sheet_name(f"{scheme_code}_{scheme_name}", max_len=31)
            out_df = merged[['date', 'Fund Start', 'Fund End', 'Fund % Change', 'NIFTY Start', 'NIFTY End', 'NIFTY % Change']].copy()
            # friendly column names for Excel
            out_df.columns = ['date', 'Fund Start', 'Fund End', 'Fund % Change', 'NIFTY Start', 'NIFTY End', 'NIFTY % Change']
            out_df.to_excel(writer, index=False, sheet_name=sheet_name)
            print(f"[+] Wrote sheet: {sheet_name} (rows={len(out_df)})")

        except Exception as e:
            print(f"[!] Error processing {scheme_name} ({scheme_code}): {e}")
            continue

    # ---------- Write Summary ----------
    if summary_rows:
        summary_df = pd.DataFrame(summary_rows)
        cols = [
            "Query", "Matched Scheme", "Scheme Code", "Fund House", "Data Points", "Correlation",
            "With Market %", "Avg Fund Return (%)", "Avg Nifty Return (%)",
            "Up Capture (%)", "Down Capture (%)", "Behavior", "Market Tolerance"
        ]
        summary_df = summary_df[cols]
        # write Summary (will reorder to first later)
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        wrote_any_sheet = True
    else:
        pd.DataFrame({"Message": ["No valid fund data found."]}).to_excel(
            writer, index=False, sheet_name="No_Data"
        )
        wrote_any_sheet = False

    # Safely close workbook
    try:
        writer.close()
    except Exception as e:
        print(f"[!] Warning while closing Excel writer: {e}")

    # Apply conditional formatting if Summary exists, and ensure it is first sheet
    try:
        wb = load_workbook(OUTPUT_FILE)
        if "Summary" in wb.sheetnames:
            # move Summary to first position
            summary_sheet = wb["Summary"]
            wb._sheets.remove(summary_sheet)
            wb._sheets.insert(0, summary_sheet)

            ws = wb["Summary"]
            header = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
            mt_col = header.get("Market Tolerance")
            beh_col = header.get("Behavior")

            if mt_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=mt_col)
                    val = (cell.value or "").strip()
                    if val == "High":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    elif val == "Medium":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                    elif val == "Low":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

            if beh_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=beh_col)
                    val = (cell.value or "").strip()
                    if val == "Against Market":
                        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Blue
        wb.save(OUTPUT_FILE)
        print(f"\n✅ Done. Excel saved to: {OUTPUT_FILE}")
    except Exception as e:
        print(f"[!] Formatting skipped: {e}")
        wb.save(OUTPUT_FILE)
        print(f"✅ Data saved (no formatting): {OUTPUT_FILE}")

if __name__ == "__main__":
    analyze_funds(FUND_QUERIES)